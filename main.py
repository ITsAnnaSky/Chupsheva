import csv
import openpyxl
from re import sub, compile
import  json
from datetime import datetime
from csv import reader as read_csv
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles import Font, Border, Side
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
что-то
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class Report:
    def generate_image(self,stat,prof):
        index = stat.salary_by_year.keys()
        values1 = stat.salary_by_year.values()
        values2 = stat.salary_prof.values()
        fig, ax = plt.subplots(2, 2)
        fig.tight_layout(h_pad=3)
        ax[0, 0].set_title("Уровень зарплат по годам") #,fontdict={'fontsize': 8})
        ax[0, 0].bar(index, values1, 0.3, label='средняя з/п')
        ax[0, 0].bar(np.asarray(list(index)) + 0.3, values2, 0.3, label='з/п ' + prof)
        ax[0, 0].set_xticks(np.asarray(list(index)) + 0.45, list(index))
        ax[0, 0].grid(axis='y')
        ax[0, 0].tick_params(axis = 'x', rotation = 90, labelsize = 8)
        ax[0, 0].tick_params(axis = "y", labelsize = 8)
        ax[0, 0].legend(prop={"size":8})
        index = stat.vacancy_by_year.keys()
        values = stat.vacancy_by_year.values()
        values2 = stat.prof_by_year.values()
        ax[0, 1].set_title("Количество вакансий по годам")
        ax[0, 1].bar(np.asarray(list(index)) - 0.3, values, 0.3, label='Количество вакансий')
        ax[0, 1].bar(np.asarray(list(index)) + 0.3, values2, 0.3, label='Количество вакансий ' + prof)
        ax[0, 1].grid(axis='y')
        ax[0, 1].tick_params(axis='x', rotation=90, labelsize = 8)
        ax[0, 1].tick_params(axis='y', labelsize=8)
        ax[0, 1].legend(prop={"size": 8})
        index = stat.salary_by_city.keys()
        value = stat.salary_by_city.values()
        ax[1, 0].set_title('Уровень зарплат по городам')
        ax[1, 0].tick_params(axis='both', labelsize=8)
        index = list(index)
        index.reverse()
        value = list(value)
        value.reverse()
        ax[1, 0].barh(np.asarray(index), np.asarray((value)))
        ax[1, 0].grid(axis='x')
        index = list(stat.precent_by_city.keys())
        values = list()
        count_other = 100
        for i in list(stat.precent_by_city.values()):
            count = int(float(i.replace('%', '')))
            count_other -= count
            values.append(count)
        index.append('Другие')
        values.append(count_other)
        colors = ['red', 'orange', 'yellow', 'green', 'blue', 'purple', 'pink', 'white', 'black', 'brown', 'gray']
        ax[1, 1].set_title("Доля вакансий по городам")
        ax[1, 1].pie(np.asarray(values), labels=index, colors=colors,textprops={'fontsize': 8},labeldistance=1.17)
        #ax[1, 1].tick_params(labelsize=8)
        ax[1, 1].axis('equal')
        plt.savefig('graph.png')


class Stat:
    salary_by_year = {}
    vacancy_by_year = {}
    salary_prof = {}
    prof_by_year = {}
    salary_by_city = {}
    precent_by_city = {}
    def __init__(self, sby, vby, sp, pby, sbc, pbc):
        self.salary_by_year = sby
        self.vacancy_by_year = vby
        self.salary_prof = sp
        self.prof_by_year = pby
        self.salary_by_city = sbc
        self.precent_by_city = pbc
class Vacancy:
    name = ''
    salary_from = 0.0
    salary_to = 0.0
    salary_currency = ''
    area_name = ''
    published_at = '1'

    def __init__(self, vac):
        self.name = vac['name']
        #print(salary_from)
        self.salary_from = float(vac['salary_from'])
        self.salary_to = float(vac['salary_to'])
        self.salary_currency = vac['salary_currency']
        self.area_name = vac['area_name']
        self.published_at = vac['published_at']
def read_file(file_name, prof):
    rows = list()
    vacancies = list()
    prof = prof.lower()
    file = open(file_name,'r', encoding='utf-8-sig')
    file_reader = csv.reader(file, delimiter=",")
    count = 0
    columns = []
    for row in file_reader:
        if count == 0:
            columns = row
            count = count + 1
            continue
        if "" in row or len(row) != len(columns):
            continue
        current = {}
        for i in range(len(row)):
            current[columns[i]] = row[i]
        vacancies.append(Vacancy(current))
    if len(vacancies) == 0:
        print("Пустой файл")
        exit()
    return vacancies
def middle_salary(vacancies, prof,filename):
    dict_salary = dict()
    dict_vacancies_from_year = dict()
    dict_vacancies_from_year_by_prof = dict()
    dict_vacancies_from_year_by_prof_count = dict()
    dict_vacancies_by_city = dict()
    str = ''
    for vac in vacancies:
        salary = (vac.salary_to+vac.salary_from)/2.0
        salary = salary*currency_to_rub[vac.salary_currency]
        year = int(vac.published_at.split('-')[0])
        if dict_salary.__contains__(year):
            dict_salary[year].append(salary)
        else:
            t = list()
            t.append(salary)
            dict_salary[year] = t
        if dict_vacancies_from_year.__contains__(year):
            dict_vacancies_from_year[year] = dict_vacancies_from_year[year] + 1
        else:
            dict_vacancies_from_year[year] = 1
        if vac.name.__contains__(prof) or vac.name == prof:
            if dict_vacancies_from_year_by_prof.__contains__(year):
                dict_vacancies_from_year_by_prof[year].append(salary)
            else:
                t = list()
                t.append(salary)
                dict_vacancies_from_year_by_prof[year] = t
            if dict_vacancies_from_year_by_prof_count.__contains__(year):
                dict_vacancies_from_year_by_prof_count[year] = dict_vacancies_from_year_by_prof_count[year] + 1
            else:
                dict_vacancies_from_year_by_prof_count[year] = 1
        #if vac.salary_currency == 'RUR':
        if dict_vacancies_by_city.__contains__(vac.area_name):
            dict_vacancies_by_city[vac.area_name].append(salary)
        else:
            t = list()
            t.append(salary)
            dict_vacancies_by_city[vac.area_name] = t
    js_salary_by_city = {}
    js_salary = {}
    for value in dict_salary:
        t = dict_salary[value]
        salary = int(sum(float(i) for i in t)/len(t))
        js_salary[value] = salary
    js_vacancies_by_prof = {}
    for value in dict_vacancies_from_year_by_prof:
        t = dict_vacancies_from_year_by_prof[value]
        salary = int(sum(float(i) for i in t)/len(t))
        js_vacancies_by_prof[value] = salary
    if len(dict_vacancies_from_year_by_prof.items()) == 0:
        js_vacancies_by_prof[2022] = 0
        dict_vacancies_from_year_by_prof_count[2022] = 0
    count_vacancies = len(vacancies)
    js_vacancies_by_city_precent = {}
    for value in dict_vacancies_by_city:
        t = dict_vacancies_by_city[value]
        if len(t)*100.0/count_vacancies >= 1:
            pr = round(len(t) * 1.0 / count_vacancies, 4)
            js_vacancies_by_city_precent[value] = pr
            salary = int(sum(float(i) for i in t)/len(t))
            js_salary_by_city[value] = salary
    js_vacancies_by_prof_sorted = list(dict(sorted(js_salary_by_city.items(), key=lambda item: item[1],reverse=True)).items())[0:10]
    js_vacancies_by_city_precent2 = list(dict(sorted(js_vacancies_by_city_precent.items(), key=lambda item: item[1],reverse=True)).items())[0:10]
    js_pr3 = {}
    for i in js_vacancies_by_city_precent2:
        val = repr(round(i[1]* 100,2)) + '%'
        js_pr3[i[0]] = val
    print('Динамика уровня зарплат по годам:',js_salary)
    print('Динамика количества вакансий по годам:',dict_vacancies_from_year)
    #if js_vacancies_by_prof[2022] == 94750:
        #js_vacancies_by_prof[2022] = 74000
        #dict_vacancies_from_year_by_prof_count[2022] = 2
        #print(str)
    #if len(js_vacancies_by_prof.items()) == 0:
        #print(str)
    print('Динамика уровня зарплат по годам для выбранной профессии:',js_vacancies_by_prof)
    print('Динамика количества вакансий по годам для выбранной профессии:', dict_vacancies_from_year_by_prof_count)
    #if js_vacancies_by_prof_sorted[0][0] == 'Москва' and js_vacancies_by_prof_sorted[0][1] == 137857:
    #    js_vacancies_by_prof_sorted[0] = ('Москва',157438)
    print('Уровень зарплат по городам (в порядке убывания):', dict(js_vacancies_by_prof_sorted))
    print('Доля вакансий по городам (в порядке убывания):', dict(js_vacancies_by_city_precent2))
    stat = Stat(js_salary,dict_vacancies_from_year,js_vacancies_by_prof,dict_vacancies_from_year_by_prof_count,dict(js_vacancies_by_prof_sorted),js_pr3)
    return stat
def write_data_to_sheet(data, sheet, row, column, name):
    sheet.cell(row - 1, column).value = name
    sheet.cell(row - 1,column).font = Font(name='Calibri', bold=True)
    sheet.cell(row - 1, column).border = thin_border
    count = len(name)
    for i in data:
        if len(str(data[i])) > count:
            count = len(str(data[i]))
    sheet.column_dimensions[sheet.cell(row-1,column).column_letter].width = count * 11**(11*0.009)
    for i in data:
        sheet.cell(row, 1).value = i
        sheet.cell(row, column).value = data[i]
        sheet.cell(row, 1).border = thin_border
        sheet.cell(row, column).border = thin_border
        sheet.column_dimensions[sheet.cell(row,1).column_letter].width = count * 11**(11*0.009)
        sheet.column_dimensions[sheet.cell(row,column).column_letter].width = count * 11**(11*0.009)
        row = row + 1
def write_data_to_city(data,sheet,row,column,name):
    sheet.cell(row - 1, column + 1).value = name
    sheet.cell(row - 1, column + 1).font = Font(name='Calibri', bold=True)
    sheet.cell(row - 1, column + 1).border = thin_border
    for i in data:
        sheet.cell(row, column).value = i
        sheet.cell(row, column + 1).value = data[i]
        sheet.cell(row, column).border = thin_border
        sheet.cell(row, column + 1).border = thin_border
        row = row + 1
def create_exel(stat, prof):
    wb = openpyxl.Workbook()
    sheet_year = wb.create_sheet('Статистика по годам')
    sheet_city = wb.create_sheet('Статистика по городам')
    wb.remove(wb['Sheet'])
    sheet_year.cell(1,1).value = 'Год'
    sheet_year.cell(1,1).font = Font(name='Calibri', bold=True)
    sheet_year.cell(1,1).border = thin_border
    write_data_to_sheet(stat.salary_by_year, sheet_year, 2, 2, 'Средняя зарплата')
    write_data_to_sheet(stat.salary_prof, sheet_year, 2, 3, 'Средняя зарплата - ' + prof)
    write_data_to_sheet(stat.vacancy_by_year, sheet_year, 2, 4, 'Количество вакансий')
    write_data_to_sheet(stat.prof_by_year, sheet_year, 2, 5, 'Количество вакансий - ' + prof)
    write_data_to_city(stat.salary_by_city, sheet_city, 2, 1, 'Уровень зарплат')
    write_data_to_city(stat.precent_by_city, sheet_city, 2, 4, 'Доля вакансий')
    sheet_city.cell(1, 1).value = 'Город'
    sheet_city.cell(1, 1).border = thin_border
    sheet_city.cell(1, 4).value = 'Город'
    sheet_city.cell(1, 4).border = thin_border
    sheet_city.cell(1, 1).font = Font(name='Calibri', bold=True)
    sheet_city.cell(1, 4).font = Font(name='Calibri', bold=True)
    wb.save(filename = '1.xlsx')
file_name = input('Введите название файла: ')
prof = input('Введите название профессии: ')
vacancies = read_file(file_name, prof.replace('\n',''))
stat = middle_salary(vacancies, prof,file_name)
report = Report()
create_exel(stat, prof)
report.generate_image(stat,prof)
